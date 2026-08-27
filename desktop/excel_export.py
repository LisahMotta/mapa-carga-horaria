"""
Geração do Mapa de Carga Horária em Excel (.xlsx), reproduzindo o layout do
formulário oficial ANEXO III (SEDUC / CGRH / URE São José dos Campos).

Depende de openpyxl. Mantido separado da interface para facilitar testes.
"""

from __future__ import annotations

from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from calculo import MESES, JORNADAS

# ------------------------------------------------------------------ estilos
_THIN = Side(style="thin", color="000000")
_MED = Side(style="medium", color="000000")
BORDA = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BORDA_MED = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

F_TITULO = Font(name="Arial", size=11, bold=True)
F_SUB = Font(name="Arial", size=9, bold=True)
F_NUM = Font(name="Arial", size=8, bold=True)
F_TXT = Font(name="Arial", size=9)
F_TXT_B = Font(name="Arial", size=9, bold=True)
F_OBS = Font(name="Arial", size=8)
F_BIG = Font(name="Arial", size=14, bold=True)

CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTRO_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)
ESQ = Alignment(horizontal="left", vertical="center", wrap_text=True)
ESQ_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

FILL_HDR = PatternFill("solid", fgColor="D9E1F2")
FILL_TOT = PatternFill("solid", fgColor="E2EFDA")

OBS_TEXTO = [
    "A.  TITULAR DE CARGO",
    "A1 - Elaborar quadro de carga horária dos últimos 60 meses, preencher campo",
    "(7) discriminando mensalmente o somatório da Jornada + Carga Suplementar.",
    "A2 - elaborar quadro de carga horária de acordo com opção, por período de (84)",
    "meses ininterruptos ou (120) meses intercalados preencher campo (7) discrimi-",
    "nando mensalmente o somatório da Jornada + Carga Suplementar no período de",
    "opção do docente.",
    "B.  OCUPANTE DE FUNÇÃO ATIVIDADE - OFA",
    "Preencher campo (7) discriminando mensalmente a C.H. exercida nos últimos",
    "60 meses.",
    "C. Para períodos anteriores a 01/02/98, efetuar a equivalência entre horas e",
    "horas-aula - Consultar ANEXO I.",
    "D. Período anterior a 01/02/98 em regime de 40 horas não aplica equivalência.",
    "Verificar sempre a SED para evitar divergências. Homologação: URE São José dos Campos.",
]


def _cell(ws, coord, value=None, font=F_TXT, align=ESQ, border=BORDA, fill=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = font
    c.alignment = align
    if border is not None:
        c.border = border
    if fill is not None:
        c.fill = fill
    return c


def _merge(ws, rng):
    ws.merge_cells(rng)


def _bordar_intervalo(ws, r1, c1, r2, c2, border=BORDA):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def _marca(cond: bool) -> str:
    return "X" if cond else " "


def gerar_mapa_excel(dados: Dict, caminho: str) -> None:
    """Gera o arquivo .xlsx do Mapa de Carga Horária em `caminho`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Carga Horária"
    ws.sheet_view.showGridLines = False

    meses: List[Tuple[int, int]] = dados["meses"]
    anos = sorted({a for a, _ in meses})
    ativos = {(a, m) for a, m in meses}
    ch = dados["ch"]

    # Colunas: A=rótulo nº, B=meses, C.. = anos
    col_ano0 = 3  # C
    ult_col_ano = col_ano0 + len(anos) - 1
    # bloco de observações começa depois da tabela (mín. coluna M=13, como no modelo)
    obs_c1 = max(13, ult_col_ano + 2)
    obs_c2 = obs_c1 + 7
    total_cols = max(20, obs_c2)  # T = 20

    L = get_column_letter

    # ---- larguras
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 13
    for c in range(col_ano0, ult_col_ano + 1):
        ws.column_dimensions[L(c)].width = 8
    for c in range(ult_col_ano + 1, total_cols + 1):
        if ws.column_dimensions[L(c)].width is None:
            ws.column_dimensions[L(c)].width = 9

    tl = L(total_cols)  # última coluna (normalmente T)

    # ---- linha 1: título
    _merge(ws, f"A1:{tl}1")
    _cell(ws, "A1", "ANEXO III  —  QUADRO DA CARGA HORÁRIA PARA CÁLCULO DE PROVENTOS",
          font=F_TITULO, align=CENTRO, border=BORDA_MED)
    ws.row_dimensions[1].height = 22

    # ---- linha 2: subtítulo
    _merge(ws, f"A2:{tl}2")
    _cell(ws, "A2", "SEDUC  /  CGRH  /  URE SÃO JOSÉ DOS CAMPOS",
          font=F_SUB, align=CENTRO)
    ws.row_dimensions[2].height = 16

    # ---- linha 3: nome / rg-cpf
    _cell(ws, "A3", 1, font=F_NUM, align=CENTRO)
    _merge(ws, "B3:H3")
    _cell(ws, "B3", f"NOME: {dados['nome']}", font=F_TXT_B, align=ESQ)
    _cell(ws, "I3", 2, font=F_NUM, align=CENTRO)
    _merge(ws, f"J3:{tl}3")
    _cell(ws, "J3", f"RG. {dados['rg']}      CPF. {dados['cpf']}", font=F_TXT, align=ESQ)
    ws.row_dimensions[3].height = 16

    # ---- linha 4: cargo / faixa-di / vínculo
    _cell(ws, "A4", 3, font=F_NUM, align=CENTRO)
    _merge(ws, "B4:G4")
    _cell(ws, "B4", f"CARGO/FUNÇÃO-ATIVIDADE: {dados['cargo']}", font=F_TXT, align=ESQ)
    _cell(ws, "H4", 4, font=F_NUM, align=CENTRO)
    _merge(ws, "I4:K4")
    _cell(ws, "I4", f"FAIXA/NÍVEL: {dados['faixa']}   DI: {dados['di']}", font=F_TXT, align=ESQ)
    _cell(ws, "L4", 5, font=F_NUM, align=CENTRO)
    _merge(ws, f"M4:{tl}4")
    titular = dados["vinculo"] == "titular"
    _cell(ws, "M4",
          f"( {_marca(titular)} ) TITULAR DE CARGO        "
          f"( {_marca(not titular)} ) OCUPANTE DE FUNÇÃO ATIVIDADE",
          font=F_TXT, align=ESQ)
    ws.row_dimensions[4].height = 16

    # ---- linhas 5-6: jornada
    _merge(ws, "A5:A6")
    _cell(ws, "A5", 6, font=F_NUM, align=CENTRO)
    j = int(dados["jornada"])
    tabela = dados.get("jornada_tabela", "")
    # Linha "Preencher Quadro": jornadas da carreira selecionada, com X na escolhida
    jornadas_carr = dados.get("jornadas_carreira", [])
    partes = "    ".join(
        f"( {_marca(int(jc['tabela']) == int(tabela))} ) T{jc['tabela']} Jornada {jc['nome']}"
        for jc in jornadas_carr
    )
    _merge(ws, f"B5:{tl}5")
    _cell(ws, "B5", f"Preencher Quadro:   {partes}", font=F_TXT, align=ESQ)
    _merge(ws, f"B6:{tl}6")
    _cell(ws, "B6",
          f"Titular de Cargo atualmente incluído na Jornada {dados['jornada_nome']}, "
          f"a partir de {dados['desde']}, DOE {dados['doe']}  ( {j} horas = Tabela {tabela} )",
          font=F_TXT, align=ESQ)
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 16

    # ---- linha 7: cabeçalhos CARGA HORÁRIA / OBSERVAÇÕES
    _cell(ws, "A7", 7, font=F_NUM, align=CENTRO)
    _merge(ws, f"B7:{L(ult_col_ano)}7")
    _cell(ws, "B7", "CARGA HORÁRIA", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)
    _merge(ws, f"{L(obs_c1)}7:{L(obs_c2)}7")
    _cell(ws, f"{L(obs_c1)}7", "OBSERVAÇÕES", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)

    # ---- linha 8-9: ANO / MÊS + anos
    _cell(ws, "B8", "ANO", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)
    _cell(ws, "B9", "MÊS", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)
    for j_idx, ano in enumerate(anos):
        col = col_ano0 + j_idx
        _merge(ws, f"{L(col)}8:{L(col)}9")
        _cell(ws, f"{L(col)}8", ano, font=F_TXT_B, align=CENTRO, fill=FILL_HDR)

    # bloco de observações (col obs_c1..obs_c2), linhas 8..21
    for i, txt in enumerate(OBS_TEXTO):
        linha = 8 + i
        _merge(ws, f"{L(obs_c1)}{linha}:{L(obs_c2)}{linha}")
        _cell(ws, f"{L(obs_c1)}{linha}", txt, font=F_OBS, align=ESQ, border=None)

    # ---- linhas 10-21: meses e valores
    for i in range(12):
        linha = 10 + i
        _cell(ws, f"B{linha}", MESES[i].upper(), font=F_TXT, align=ESQ)
        for j_idx, ano in enumerate(anos):
            col = col_ano0 + j_idx
            key = (ano, i)
            val = ""
            if key in ativos and ch.get(f"{ano}-{i + 1:02d}") is not None:
                val = int(round(float(ch[f"{ano}-{i + 1:02d}"])))
            _cell(ws, f"{L(col)}{linha}", val if val != "" else None,
                  font=F_TXT, align=CENTRO)
        ws.row_dimensions[linha].height = 14

    # ---- linhas 22-23: TOTAIS ANUAIS
    _cell(ws, "B22", "TOTAIS", font=F_TXT_B, align=CENTRO, fill=FILL_TOT)
    _cell(ws, "B23", "ANUAIS", font=F_TXT_B, align=CENTRO, fill=FILL_TOT)
    for j_idx, ano in enumerate(anos):
        col = col_ano0 + j_idx
        _merge(ws, f"{L(col)}22:{L(col)}23")
        formula = f"=SUM({L(col)}10:{L(col)}21)"
        _cell(ws, f"{L(col)}22", formula, font=F_TXT_B, align=CENTRO, fill=FILL_TOT)

    # ---- linha 24: 7.A total geral
    _cell(ws, "J24", "7.A", font=F_NUM, align=CENTRO)
    _merge(ws, f"K24:{L(total_cols - 1)}24")
    _cell(ws, "K24", "Some os totais anuais - TOTAL GERAL DA CARGA HORÁRIA EM HORAS  =",
          font=F_TXT_B, align=ESQ)
    somas = "+".join(f"{L(col_ano0 + i)}22" for i in range(len(anos)))
    _cell(ws, f"{tl}24", f"={somas}" if somas else 0, font=F_TXT_B, align=CENTRO)
    ws.row_dimensions[24].height = 16

    # ---- linha 25: campo 8 label / declaração / assinatura superior
    _cell(ws, "A25", 8, font=F_NUM, align=CENTRO)
    _merge(ws, "B25:I25")
    _cell(ws, "B25", "Nomeação/Designação em regime de 40 horas/semanais:",
          font=F_TXT_B, align=ESQ)
    _cell(ws, "J25", 9, font=F_NUM, align=CENTRO)
    _merge(ws, "K25:P25")
    _cell(ws, "K25", "DECLARAÇÃO", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)
    _merge(ws, f"Q25:{tl}25")
    _cell(ws, "Q25", "ASSINATURA DO SUPERIOR IMEDIATO", font=F_NUM, align=CENTRO)

    # nomeação (campo 8): A26:I31
    _merge(ws, "A26:I31")
    _cell(ws, "A26", dados["nomeacao"] or "Não há.", font=F_TXT, align=ESQ_TOP)

    # declaração (K26:P30)
    _merge(ws, "K26:P30")
    _cell(ws, "K26",
          "Declaro que estou ciente do nº de aulas constante deste Quadro, que retrata a "
          "minha opção nos termos do Art. 39 (das DDTT) da LC 836/97 "
          f"( {dados['n_meses']} meses) para fins de cálculo de proventos.",
          font=F_TXT, align=ESQ_TOP)
    _merge(ws, "K31:P31")
    _cell(ws, "K31", "Assinatura do(a) interessado(a)", font=F_OBS, align=CENTRO)
    # espaço p/ assinatura do superior (Q26:T31)
    _merge(ws, f"Q26:{tl}31")
    _cell(ws, "Q26", "", align=CENTRO)

    # ---- linhas 33-34: instruções campo 10
    _cell(ws, "J33", 10, font=F_NUM, align=CENTRO)
    _merge(ws, f"K33:{tl}33")
    _cell(ws, "K33",
          "Para determinar a MÉDIA DA CARGA HORÁRIA, divide-se o valor obtido em (7.A) pelo",
          font=F_OBS, align=ESQ, border=None)
    _merge(ws, f"K34:{tl}34")
    _cell(ws, "K34",
          "nº de meses da opção em (9), arredondando-se para o inteiro a fração resultante.",
          font=F_OBS, align=ESQ, border=None)

    # ---- linhas 35-37: quadro campo 10
    _merge(ws, "K35:N35")
    _cell(ws, "K35", "MÉDIA CARGA HORÁRIA", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)
    _merge(ws, "P35:Q35")
    _cell(ws, "P35", "JORNADA (6)", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)
    _merge(ws, f"S35:{tl}35")
    _cell(ws, "S35", "CARGA SUPLEMENTAR", font=F_TXT_B, align=CENTRO, fill=FILL_HDR)

    completo = dados["tipo_quadro"] == "completo"
    _merge(ws, "K36:N36")
    _cell(ws, "K36", dados["media"], font=F_BIG, align=CENTRO)
    _cell(ws, "O36", "=" if completo else "", font=F_TXT_B, align=CENTRO, border=None)
    _merge(ws, "P36:Q36")
    _cell(ws, "P36", dados["jornada"] if completo else "", font=F_BIG, align=CENTRO)
    _cell(ws, "R36", "+" if completo else "", font=F_TXT_B, align=CENTRO, border=None)
    _merge(ws, f"S36:{tl}36")
    _cell(ws, "S36", dados["suplementar"] if completo else "", font=F_BIG, align=CENTRO)

    _merge(ws, "K37:N37")
    _cell(ws, "K37", "TITULAR DE CARGO/OFA", font=F_OBS, align=CENTRO)
    _merge(ws, "P37:Q37")
    _cell(ws, "P37", "TITULAR DE CARGO", font=F_OBS, align=CENTRO)
    _merge(ws, f"S37:{tl}37")
    _cell(ws, "S37", "TITULAR DE CARGO", font=F_OBS, align=CENTRO)
    ws.row_dimensions[36].height = 26

    # ---- bordas nas áreas principais
    _bordar_intervalo(ws, 3, 1, 6, total_cols)
    _bordar_intervalo(ws, 7, 1, 23, ult_col_ano)     # tabela carga horária + rótulos
    _bordar_intervalo(ws, 24, 10, 24, total_cols)    # 7.A
    _bordar_intervalo(ws, 25, 1, 31, total_cols)     # campo 8/9
    _bordar_intervalo(ws, 35, 11, 37, total_cols)    # campo 10

    # ---- impressão
    ws.print_area = f"A1:{tl}39"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    wb.save(caminho)
